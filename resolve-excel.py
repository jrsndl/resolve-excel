import argparse
import logging
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side
import os
import re
import sys

def make_excel(images, edit_index, root):
    w = 160
    h = 90
    column_width = w * 0.135
    row_height = h * 0.761
    wb = Workbook()
    ws = wb.worksheets[0]
    ws.append(["Thumbnail", "Shot"])
    ws.column_dimensions['A'].width = column_width
    ws.column_dimensions['B'].width = column_width

    cnt = 1
    step = 10
    for one_image in images:
        row = cnt + 1

        # image
        ws.row_dimensions[row].height = row_height
        img = Image(one_image)
        img.width = w
        img.height = h
        img.anchor = f"A{row}"
        ws.add_image(img)

        # shot name
        ws[f"B{row}"] = f"sq10_sh{cnt*step:03d}"
        ws[f"B{row}"].font = Font(bold=True, size=16)
        ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")



        cnt += 1

    wb.save(f'{root}\out.xlsx')

def sort_images(image_list):
    """
    gallery images from Resolve are named name track.shot.version
    """
    REGEX = "^(.*)(\d{1,2})\.(\d{1,4})\.(\d{1,2})$"
    images = []
    img_dict = {}
    for image in image_list:
        name = os.path.basename(image)
        no_ext = ".".join(name.split('.')[:-1])
        if re.match(REGEX, no_ext):
            rgx = re.search(REGEX, no_ext)
            try:
                to_sort = f"{rgx.group(2).zfill(2)}.{rgx.group(3).zfill(4)}.{rgx.group(4).zfill(2)}"
                img_dict[to_sort] = image
            except:
                pass
    img_dict = dict(sorted(img_dict.items()))
    for key, val in img_dict.items():
        images.append(val)
    return images

def get_file_list(root, logger, include="", exclude="", pattern=None, recursive=False):
    if not os.path.isdir(root):
        logger("Folder path unreachable: {}".format(root))

    file_list = []
    files = []
    if recursive:
        files = [os.path.join(dirpath, f) for (dirpath, dirnames, filenames) in
                 os.walk(root) for f in filenames]
    else:
        files = os.listdir(root)
        files = [root + '/' + f for f in files if
                 os.path.isfile(root + '/' + f)]

    if files and len(files) > 0:
        for one_file in files:
            if include != '' and include not in one_file:
                continue
            if exclude != '' and exclude in one_file:
                continue
            if pattern != '':
                if not bool(re.match(pattern, os.path.basename(one_file))):
                    continue
            file_list.append(one_file)
    else:
        logger.warning("No files found at {}".format(root))
    return file_list

def get_args():
    parser = argparse.ArgumentParser(
        description="Takes images and Edit Index from Davinci Resolve, and converts to csv with thumbnails path.")
    parser.add_argument(
        '-i',
        help="Root folder for images and csv",
        type=str,
        required=False
    )
    return parser.parse_args()

def get_app_path():
    application_path = None
    if getattr(sys, 'frozen', False):
        # If the application is run as a bundle, the PyInstaller bootloader
        # extends the sys module by a flag frozen=True and sets the app
        # path into variable _MEIPASS'.
        #self.application_path = sys._MEIPASS.replace('\\', '/')
        application_path = os.path.dirname(sys.executable).replace('\\', '/')
    else:
        application_path = os.path.dirname(os.path.abspath(__file__)).replace('\\', '/')
    return application_path




if __name__ == "__main__":
    # log
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.DEBUG)
    formatter_time = logging.Formatter('%(asctime)s:%(levelname)s:%(message)s')
    formatter = logging.Formatter('%(levelname)s:%(message)s')
    app_path = get_app_path()
    if app_path is not None:
        app_path += '/thumb-spreadsheet.log'
        file_handler = logging.FileHandler(app_path)
        logger.addHandler(file_handler)
    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)

    args = vars(get_args())
    search_root = args.get('i')
    if search_root is None or len(search_root) == 0:
        logger.error("No root folder specified.")
        exit(1)
    search_root.replace('\\', '/')
    if search_root.endswith('/'):
        search_root = search_root[:-1]
    if not os.path.exists(search_root):
        logger.error("Root folder does not exist.")
        exit(1)
    if os.path.isfile(search_root):
        search_root = os.path.dirname(search_root)
    if not os.path.isdir(search_root):
        logger.error(f"Root folder {search_root} is not a directory.")
        exit(1)

    image_list = get_file_list(search_root, logger, pattern=r".*\.(jpg|jpeg|png)$")
    if image_list is None or len(image_list) == 0:
        logger.error("No images found.")
        exit(2)

    images = sort_images(image_list)
    if images is None or len(images) == 0:
        logger.error("No images found.")
        exit(2)
    logger.info(f"Found {len(images)} images.")

    edit_index = ""
    edit_indexes = get_file_list(search_root, logger, pattern=".*\.(csv)$")
    if edit_indexes is None or len(edit_indexes) == 0:
        logger.warning("No edit indexes found.")
        edit_index = ""
    elif len(edit_indexes) > 1:
        logger.warning(f"Found {len(edit_indexes)} edit indexes. Picking the first one: {edit_indexes[0]}")
    edit_index = edit_indexes[0]

    make_excel(images, edit_index, search_root)







